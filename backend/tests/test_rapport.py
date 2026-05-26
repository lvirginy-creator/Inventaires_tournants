"""Tests d'intégration — Rapport d'inventaire.

Couvre :
- Rapport JSON disponible pour en_cours / terminee / validee
- Rapport refusé pour brouillon → 409
- Calculs écarts (qt_comptee, ecart, ecart_pct)
- Tri par |écart| décroissant
- Export CSV (content-type, BOM, valeurs)
- Export XLSX (content-type, parsable)
"""

from __future__ import annotations

import io
import uuid

import openpyxl
import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import now_utc
from app.core.security import create_tablette_token, hash_jwt
from app.models.article import Article
from app.models.campagne import Campagne, StatutCampagne
from app.models.comptage import Comptage
from app.models.magasin import Magasin
from app.models.tablette import RoleTablette, SessionTablette, Tablette
from app.models.utilisateur import Utilisateur

# ── Helpers ────────────────────────────────────────────────────────────────────


async def _make_session(magasin_id: uuid.UUID, db: AsyncSession) -> SessionTablette:
    """Crée une tablette + session pour insérer des comptages."""
    t = Tablette(
        magasin_id=magasin_id,
        nom=f"Tab-{uuid.uuid4().hex[:6]}",
        device_id=f"dev-{uuid.uuid4().hex[:6]}",
    )
    db.add(t)
    await db.flush()

    sid = uuid.uuid4()
    tok = create_tablette_token(sid, t.id, magasin_id, "operateur")
    s = SessionTablette(
        id=sid,
        tablette_id=t.id,
        magasin_id=magasin_id,
        role=RoleTablette.operateur,
        jwt_token_hash=hash_jwt(tok),
    )
    db.add(s)
    await db.flush()
    return s


async def _add_comptage(
    campagne_id: uuid.UUID,
    article_id: uuid.UUID,
    session: SessionTablette,
    quantite: float,
    db: AsyncSession,
) -> None:
    c = Comptage(
        campagne_id=campagne_id,
        article_id=article_id,
        magasin_id=session.magasin_id,
        session_id=session.id,
        quantite=quantite,
        client_uuid=str(uuid.uuid4()),
        counted_at=now_utc(),
    )
    db.add(c)
    await db.flush()


# ── Tests JSON ─────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_rapport_brouillon_interdit(
    client: AsyncClient,
    auth_headers: dict,
    campagne: Campagne,
):
    """Statut brouillon → 409."""
    assert campagne.statut == StatutCampagne.brouillon
    resp = await client.get(
        f"/api/v1/campagnes/{campagne.id}/rapport",
        headers=auth_headers,
    )
    assert resp.status_code == 409


@pytest.mark.asyncio
async def test_rapport_en_cours(
    client: AsyncClient,
    auth_headers: dict,
    campagne: Campagne,
    article: Article,
    magasin: Magasin,
    db: AsyncSession,
):
    """Rapport disponible dès qu'une campagne est en cours."""
    await client.post(
        f"/api/v1/campagnes/{campagne.id}/articles",
        json={"article_id": str(article.id), "quantite_theorique": 20},
        headers=auth_headers,
    )
    await client.post(f"/api/v1/campagnes/{campagne.id}/demarrer", headers=auth_headers)

    resp = await client.get(
        f"/api/v1/campagnes/{campagne.id}/rapport",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["statut"] == "en_cours"
    assert data["nb_articles"] == 1
    assert len(data["lignes"]) == 1


@pytest.mark.asyncio
async def test_rapport_calcul_ecarts(
    client: AsyncClient,
    auth_headers: dict,
    magasin: Magasin,
    admin_user: Utilisateur,
    db: AsyncSession,
    societe,
):
    """Vérifie le calcul précis des écarts et du tri."""
    # 3 articles avec des écarts différents
    arts = []
    for i, (cb, qt_theo) in enumerate([("CB-A", 10), ("CB-B", 100), ("CB-C", 50)]):
        a = Article(
            societe_id=societe.id,
            code_barre=f"{cb}-{uuid.uuid4().hex[:4]}",
            code_article=f"ART-{i}",
            libelle=f"Article {i}",
        )
        db.add(a)
        arts.append((a, qt_theo))
    await db.flush()

    # Campagne
    from app.models.campagne import Campagne, LigneCampagne

    camp = Campagne(magasin_id=magasin.id, nom="Camp-Ecarts", created_by=admin_user.id)
    db.add(camp)
    await db.flush()

    for art, qt_theo in arts:
        db.add(LigneCampagne(campagne_id=camp.id, article_id=art.id, quantite_theorique=qt_theo))
    await db.flush()

    # Démarrer
    await client.post(f"/api/v1/campagnes/{camp.id}/demarrer", headers=auth_headers)

    # Comptages : art0→7 (écart=-3), art1→80 (écart=-20), art2→50 (écart=0)
    sess = await _make_session(magasin.id, db)
    await _add_comptage(camp.id, arts[0][0].id, sess, 7, db)
    await _add_comptage(camp.id, arts[1][0].id, sess, 80, db)
    await _add_comptage(camp.id, arts[2][0].id, sess, 50, db)

    resp = await client.get(
        f"/api/v1/campagnes/{camp.id}/rapport",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    data = resp.json()

    assert data["nb_articles"] == 3
    assert data["nb_articles_comptes"] == 3
    assert data["nb_articles_ok"] == 1  # art2
    assert data["nb_articles_en_ecart"] == 2  # art0 (-3) et art1 (-20)

    # Tri : |écart| décroissant → art1(-20) avant art0(-3) avant art2(0)
    lignes = data["lignes"]
    ecarts = [float(lg["ecart"]) for lg in lignes]
    assert ecarts[0] == -20.0
    assert ecarts[1] == -3.0
    assert ecarts[2] == 0.0

    # ecart_pct art1 : -20/100 * 100 = -20.0 %
    assert abs(float(lignes[0]["ecart_pct"]) - (-20.0)) < 0.01
    # ecart_pct art0 : -3/10 * 100 = -30.0 %
    assert abs(float(lignes[1]["ecart_pct"]) - (-30.0)) < 0.01


@pytest.mark.asyncio
async def test_rapport_article_sans_comptage(
    client: AsyncClient,
    auth_headers: dict,
    campagne: Campagne,
    article: Article,
    db: AsyncSession,
):
    """Article dans la campagne mais pas de comptage → qt_comptee=0, ecart=-theo."""
    await client.post(
        f"/api/v1/campagnes/{campagne.id}/articles",
        json={"article_id": str(article.id), "quantite_theorique": 15},
        headers=auth_headers,
    )
    await client.post(f"/api/v1/campagnes/{campagne.id}/demarrer", headers=auth_headers)

    resp = await client.get(
        f"/api/v1/campagnes/{campagne.id}/rapport",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    ligne = resp.json()["lignes"][0]
    assert float(ligne["quantite_comptee"]) == 0.0
    assert float(ligne["ecart"]) == -15.0


@pytest.mark.asyncio
async def test_rapport_article_sans_theorique(
    client: AsyncClient,
    auth_headers: dict,
    campagne: Campagne,
    article: Article,
    db: AsyncSession,
):
    """quantite_theorique None → ecart et ecart_pct null."""
    await client.post(
        f"/api/v1/campagnes/{campagne.id}/articles",
        json={"article_id": str(article.id)},  # pas de quantite_theorique
        headers=auth_headers,
    )
    await client.post(f"/api/v1/campagnes/{campagne.id}/demarrer", headers=auth_headers)

    resp = await client.get(
        f"/api/v1/campagnes/{campagne.id}/rapport",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    ligne = resp.json()["lignes"][0]
    assert ligne["quantite_theorique"] is None
    assert ligne["ecart"] is None
    assert ligne["ecart_pct"] is None


@pytest.mark.asyncio
async def test_rapport_auth_required(client: AsyncClient, campagne: Campagne):
    resp = await client.get(f"/api/v1/campagnes/{campagne.id}/rapport")
    assert resp.status_code in (401, 403)


# ── Tests export ───────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_csv(
    client: AsyncClient,
    auth_headers: dict,
    campagne: Campagne,
    article: Article,
    db: AsyncSession,
):
    """Export CSV : content-type correct, BOM présent, données dans le fichier."""
    await client.post(
        f"/api/v1/campagnes/{campagne.id}/articles",
        json={"article_id": str(article.id), "quantite_theorique": 5},
        headers=auth_headers,
    )
    await client.post(f"/api/v1/campagnes/{campagne.id}/demarrer", headers=auth_headers)

    resp = await client.get(
        f"/api/v1/campagnes/{campagne.id}/rapport/export?format=csv",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    # BOM UTF-8
    assert resp.content[:3] == b"\xef\xbb\xbf"
    text = resp.content.decode("utf-8-sig")
    assert article.code_barre in text


@pytest.mark.asyncio
async def test_export_xlsx(
    client: AsyncClient,
    auth_headers: dict,
    campagne: Campagne,
    article: Article,
    db: AsyncSession,
):
    """Export XLSX : content-type correct, fichier parsable par openpyxl."""
    await client.post(
        f"/api/v1/campagnes/{campagne.id}/articles",
        json={"article_id": str(article.id), "quantite_theorique": 8},
        headers=auth_headers,
    )
    await client.post(f"/api/v1/campagnes/{campagne.id}/demarrer", headers=auth_headers)

    resp = await client.get(
        f"/api/v1/campagnes/{campagne.id}/rapport/export?format=xlsx",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    # Vérifier que le fichier est un XLSX valide
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws is not None
    # Ligne 1 = en-tête, ligne 2 = données
    assert ws.max_row >= 2
    # Code barre dans la première colonne de données
    assert ws.cell(row=2, column=1).value == article.code_barre


@pytest.mark.asyncio
async def test_export_brouillon_interdit(
    client: AsyncClient,
    auth_headers: dict,
    campagne: Campagne,
):
    """Export sur brouillon → 409."""
    resp = await client.get(
        f"/api/v1/campagnes/{campagne.id}/rapport/export",
        headers=auth_headers,
    )
    assert resp.status_code == 409
