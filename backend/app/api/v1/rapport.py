"""Rapport d'inventaire — résultats de comptage et écarts.

Routes :
    GET /campagnes/{campagne_id}/rapport            → JSON
    GET /campagnes/{campagne_id}/rapport/export     → CSV ou XLSX (query ?format=csv|xlsx)
"""

from __future__ import annotations

import csv
import io
import uuid
from decimal import ROUND_HALF_UP, Decimal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_admin
from app.core.database import get_db
from app.models.article import Article
from app.models.campagne import Campagne, LigneCampagne, StatutCampagne
from app.models.comptage import Comptage
from app.models.utilisateur import Utilisateur
from app.schemas.rapport import CampagneRapport, RapportLigne

router = APIRouter(prefix="/campagnes", tags=["rapport"])

# Statuts pour lesquels le rapport a du sens
_STATUTS_RAPPORT = {StatutCampagne.en_cours, StatutCampagne.terminee, StatutCampagne.validee}

_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── Helpers ────────────────────────────────────────────────────────────────────


async def _get_campagne_or_404(campagne_id: uuid.UUID, db: AsyncSession) -> Campagne:
    result = await db.execute(select(Campagne).where(Campagne.id == campagne_id))
    campagne = result.scalar_one_or_none()
    if not campagne:
        raise HTTPException(status_code=404, detail="Campagne introuvable")
    return campagne


async def _build_rapport(campagne: Campagne, db: AsyncSession) -> CampagneRapport:
    """Construit le rapport en une seule requête agrégée."""
    stmt = (
        select(
            LigneCampagne.article_id,
            Article.code_barre,
            Article.code_article,
            Article.libelle,
            Article.unite,
            LigneCampagne.quantite_theorique,
            func.coalesce(func.sum(Comptage.quantite), 0).label("quantite_comptee"),
        )
        .join(Article, Article.id == LigneCampagne.article_id)
        .outerjoin(
            Comptage,
            (Comptage.article_id == LigneCampagne.article_id)
            & (Comptage.campagne_id == LigneCampagne.campagne_id),
        )
        .where(LigneCampagne.campagne_id == campagne.id)
        .group_by(
            LigneCampagne.article_id,
            Article.code_barre,
            Article.code_article,
            Article.libelle,
            Article.unite,
            LigneCampagne.quantite_theorique,
        )
    )
    rows = (await db.execute(stmt)).all()

    lignes: list[RapportLigne] = []
    for row in rows:
        raw_theo = row.quantite_theorique
        qt_theo = Decimal(str(raw_theo)) if raw_theo is not None else None
        qt_compte = Decimal(str(row.quantite_comptee))
        if qt_theo is not None:
            ecart = (qt_compte - qt_theo).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
            ecart_pct = (float(ecart) / float(qt_theo) * 100) if qt_theo != 0 else None
        else:
            ecart = None
            ecart_pct = None
        lignes.append(
            RapportLigne(
                article_id=row.article_id,
                code_barre=row.code_barre,
                code_article=row.code_article,
                libelle=row.libelle,
                unite=row.unite,
                quantite_theorique=qt_theo,
                quantite_comptee=qt_compte,
                ecart=ecart,
                ecart_pct=round(ecart_pct, 2) if ecart_pct is not None else None,
            )
        )

    # Trier par |écart| décroissant (plus grands écarts en premier)
    lignes.sort(
        key=lambda x: abs(x.ecart) if x.ecart is not None else Decimal(0),
        reverse=True,
    )

    nb_articles = len(lignes)
    nb_comptes = sum(1 for lg in lignes if lg.quantite_comptee > 0)
    nb_ok = sum(1 for lg in lignes if lg.ecart is not None and lg.ecart == 0)
    nb_ecart = sum(1 for lg in lignes if lg.ecart is not None and lg.ecart != 0)

    return CampagneRapport(
        campagne_id=campagne.id,
        campagne_nom=campagne.nom,
        magasin_id=campagne.magasin_id,
        statut=campagne.statut,
        nb_articles=nb_articles,
        nb_articles_comptes=nb_comptes,
        nb_articles_ok=nb_ok,
        nb_articles_en_ecart=nb_ecart,
        lignes=lignes,
    )


def _rapport_to_csv(rapport: CampagneRapport) -> bytes:
    """Génère un CSV UTF-8 avec BOM (compatible Excel)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(
        [
            "Code barre",
            "Code article",
            "Libellé",
            "Unité",
            "Qté théorique",
            "Qté comptée",
            "Écart",
            "Écart %",
        ]
    )
    for lg in rapport.lignes:
        writer.writerow(
            [
                lg.code_barre,
                lg.code_article,
                lg.libelle,
                lg.unite or "",
                str(lg.quantite_theorique) if lg.quantite_theorique is not None else "",
                str(lg.quantite_comptee),
                str(lg.ecart) if lg.ecart is not None else "",
                f"{lg.ecart_pct:.2f}" if lg.ecart_pct is not None else "",
            ]
        )
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def _rapport_to_xlsx(rapport: CampagneRapport) -> bytes:
    """Génère un XLSX avec mise en forme (en-tête coloré, lignes vertes/rouges)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Inventaire"

    # ── En-tête ──────────────────────────────────────────────────────────────
    headers = [
        "Code barre",
        "Code article",
        "Libellé",
        "Unité",
        "Qté théorique",
        "Qté comptée",
        "Écart",
        "Écart %",
    ]
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── Remplissages colorés ─────────────────────────────────────────────────
    fill_ok = PatternFill("solid", fgColor="DCFCE7")  # vert pâle
    fill_ko = PatternFill("solid", fgColor="FEE2E2")  # rouge pâle

    # ── Données ──────────────────────────────────────────────────────────────
    for lg in rapport.lignes:
        row = [
            lg.code_barre,
            lg.code_article,
            lg.libelle,
            lg.unite or "",
            float(lg.quantite_theorique) if lg.quantite_theorique is not None else None,
            float(lg.quantite_comptee),
            float(lg.ecart) if lg.ecart is not None else None,
            round(lg.ecart_pct, 2) if lg.ecart_pct is not None else None,
        ]
        ws.append(row)
        if lg.ecart is not None:
            fill = fill_ok if lg.ecart == 0 else fill_ko
            for cell in ws[ws.max_row]:
                cell.fill = fill

    # ── Largeurs de colonnes ─────────────────────────────────────────────────
    col_widths = [16, 14, 36, 8, 14, 14, 10, 10]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # ── Figer la première ligne ───────────────────────────────────────────────
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Routes ─────────────────────────────────────────────────────────────────────


@router.get("/{campagne_id}/rapport", response_model=CampagneRapport)
async def get_rapport(
    campagne_id: uuid.UUID,
    _: Utilisateur = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
) -> CampagneRapport:
    """Retourne le rapport d'inventaire JSON pour une campagne.

    Disponible pour les statuts : en_cours, terminee, validee.
    """
    campagne = await _get_campagne_or_404(campagne_id, db)
    if campagne.statut not in _STATUTS_RAPPORT:
        raise HTTPException(
            status_code=409,
            detail=f"Rapport indisponible pour le statut « {campagne.statut.value} »",
        )
    return await _build_rapport(campagne, db)


@router.get("/{campagne_id}/rapport/export")
async def export_rapport(
    campagne_id: uuid.UUID,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    _: Utilisateur = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
) -> Response:
    """Télécharge le rapport au format CSV (défaut) ou XLSX.

    ?format=csv   → fichier CSV UTF-8 BOM (compatible Excel)
    ?format=xlsx  → fichier XLSX avec mise en forme
    """
    campagne = await _get_campagne_or_404(campagne_id, db)
    if campagne.statut not in _STATUTS_RAPPORT:
        raise HTTPException(
            status_code=409,
            detail=f"Export indisponible pour le statut « {campagne.statut.value} »",
        )

    rapport = await _build_rapport(campagne, db)
    safe_nom = campagne.nom.replace(" ", "_").replace("/", "-")[:50]

    if format == "xlsx":
        content = _rapport_to_xlsx(rapport)
        media_type = _XLSX_CONTENT_TYPE
        filename = f"rapport_{safe_nom}.xlsx"
    else:
        content = _rapport_to_csv(rapport)
        media_type = "text/csv; charset=utf-8"
        filename = f"rapport_{safe_nom}.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
